import psycopg2
from psycopg2._psycopg import cursor, connection
import psycopg2.extras
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from deting.DB import _get_hg_connection

SQL_TC_MEP = """
 select "CIERREARS"::real as tc
 from "UNI_CACHE_CONVERSION_ORDENADO_DESC"
 where "UNIDAD"=(select "UNI_UNIDAD_ID" from "UNI_UNIDAD" where "CODIGO"='USDM')
 and "MONEDA"=(select "UNI_UNIDAD_ID" from "UNI_UNIDAD" where "CODIGO"='ARS')
 and "FECHA"::DATE<=%s
 limit 1"""

SQL_ARANCELES_BASE = """
with USD as(
 select "CIERREARS" as tc
 from "UNI_CACHE_CONVERSION_ORDENADO_DESC"
 where "UNIDAD"=(select "UNI_UNIDAD_ID" from "UNI_UNIDAD" where "CODIGO"='USDM')
 and "MONEDA"=(select "UNI_UNIDAD_ID" from "UNI_UNIDAD" where "CODIGO"='ARS')
 limit 1
),
OPS_PRODUCTORES as(
 select distinct movs."OPERACION" as "OPERACION"
 from "OP_MOVIMIENTO" movs
 where "CUENTACONTABLE" = (select "CTB_CUENTA_ID" from "CTB_CUENTA" where "CODIGO"='3110000001')
 and "TIPOOPERACION" = (select "OP_TIPO_OPERACION_ID" from "OP_TIPO_OPERACION" where "NOMBRE"='Conciliar diferencias de recuperos')
),
MOVS as (
 select ope."FECHA" as "concertacion",
   mov."FECHASIMPLE" as "liquidacion",
   cta."ID" as "cuenta",
   cta."DENOMINACION" as "denominacionCuenta",
   cta."TIPO" as "tipoCuenta",
   mov."COMPROBANTE" as "comprobante",
   mov."INFORMACION" as "informacion",
   tope."NOMBRE" as "tipoOperacion",
   ope."OP_OPERACION_ID" as "operacion",
   mer."CODIGO" as "mercado",
   ope."SEGMENTO" as "segmento",
   ope."SESION" as "sesion",
   ope."PARTE" as "parte",
   esp."CODIGO" as "especie",
   tit."NOMBRE" as "tipoTitulo",
   mov."CONCEPTO" as "concepto",
   uni."CODIGO" as "moneda",
   mov."CANTIDAD" as "cantidad",
   mov."VALUACION" as "valuacion"
 from "OP_MOVIMIENTO" as mov
   inner join "OP_OPERACION" as ope on mov."OPERACION"=ope."OP_OPERACION_ID"
      and ope."FECHA">=%s and ope."FECHA"<(%s + interval '1 day')
   inner join "UNI_UNIDAD" as uni on mov."UNIDAD"=uni."UNI_UNIDAD_ID"
   inner join "OP_TIPO_OPERACION" as tope on ope."TIPOOPERACION"=tope."OP_TIPO_OPERACION_ID"
   inner join "CTA_ESQUEMA" as cta on mov."CUENTA"=cta."CTA_ESQUEMA_ID"
   inner join "CTA_SUBCUENTA" as scta on mov."SUBCUENTA"=scta."CTA_SUBCUENTA_ID" AND scta."USO" = 'GRAL'
   left outer join "UNI_UNIDAD" as esp on ope."UNIDAD"=esp."UNI_UNIDAD_ID"
   left outer join "UNI_TIPO_TITULO" as tit on esp."TIPOTITULO"=tit."UNI_TIPO_TITULO_ID"
   left outer join "DOM_BUR_MERCADO" as mer on ope."MERCADO"=mer."DOM_BUR_MERCADO_ID"
 where  mov."ANULADOR" is null
 and cta."TIPO" ~'Comitente.*|Productor|Propia'
 and mov."CONCEPTO" IN ('AAE', 'AAG', 'CAG','AAN', 'MAG','CAE','CAN')
 and mov."CANTIDAD" != 0
 union
 select ope."FECHA" as "concertacion",
   mov."FECHASIMPLE" as "liquidacion",
   cta."ID" as "cuenta",
   cta."DENOMINACION" as "denominacionCuenta",
   cta."TIPO" as "tipoCuenta",
   mov."COMPROBANTE" as "comprobante",
   mov."INFORMACION" as "informacion",
   tope."NOMBRE" as "tipoOperacion",
   ope."OP_OPERACION_ID" as "operacion",
   mer."CODIGO" as "mercado",
   ope."SEGMENTO" as "segmento",
   ope."SESION" as "sesion",
   ope."PARTE" as "parte",
   esp."CODIGO" as "especie",
   tit."NOMBRE" as "tipoTitulo",
   mov."CONCEPTO" as "concepto",
   uni."CODIGO" as "moneda",
   mov."CANTIDAD" as "cantidad",
   mov."VALUACION" as "valuacion"
 from "OP_MOVIMIENTO" as mov
   inner join "OP_OPERACION" as ope on mov."OPERACION"=ope."OP_OPERACION_ID"
     and ope."FECHA">=%s and ope."FECHA"<(%s + interval '1 day')
   inner join "UNI_UNIDAD" as uni on mov."UNIDAD"=uni."UNI_UNIDAD_ID"
   inner join "OP_TIPO_OPERACION" as tope on ope."TIPOOPERACION"=tope."OP_TIPO_OPERACION_ID"
   inner join "CTA_ESQUEMA" as cta on mov."CUENTA"=cta."CTA_ESQUEMA_ID"
   inner join OPS_PRODUCTORES as oppr on ope."OP_OPERACION_ID"=oppr."OPERACION"
   left outer join "UNI_UNIDAD" as esp on ope."UNIDAD"=esp."UNI_UNIDAD_ID"
   left outer join "UNI_TIPO_TITULO" as tit on esp."TIPOTITULO"=tit."UNI_TIPO_TITULO_ID"
   left outer join "DOM_BUR_MERCADO" as mer on ope."MERCADO"=mer."DOM_BUR_MERCADO_ID"
 where  mov."ANULADOR" is null
 and cta."TIPO" ~'Productor|Agente'
 and mov."CONCEPTO" ~ '[AGCEMO][PC][GEN]'
 and mov."CANTIDAD" != 0
   
 order by 1),
movs2 as (
select  "concertacion"::date,
  "liquidacion",
  "cuenta",
  "denominacionCuenta",
  "tipoCuenta",
  "comprobante",
  "informacion",
  case
   when "tipoTitulo" is null and "sesion" = 'Caución' then 'Caución'
   when "tipoTitulo" is null and "informacion" like '%%Débito%%Desc.%%' then 'Comisión descubierto'
   when "tipoTitulo" is null and "informacion" like '%%DB%%Diferencia%%' then 'DB- DIF DE PRECIO'
   when "tipoTitulo" is null and "informacion" like '%%Débito%%' then 'Aranceles varios'
   when "tipoTitulo" is null and "informacion" like '%%Débito%%' then 'Aranceles varios'
   when "tipoTitulo" is null and "tipoOperacion" like '%%Débito%%' then 'Aranceles varios'
   when "tipoTitulo" is null and "informacion" like '%%Crédito%%DB%%' then 'Bonificaciones DB'
   when "tipoTitulo" is null and "informacion" like '%%Crédito%%' then 'Bonificaciones'
   when "tipoTitulo" is null and "informacion" like '%%Liquidación%%general%%' then 'Liquidación general'
   when "tipoTitulo" is null and "tipoOperacion" like '%%Crédito%%' and "informacion" like '%%Aranceles%%' then  'Bonificaciones DB'
   when "tipoTitulo" is null and "tipoOperacion" like '%%Crédito%%' then  'Bonificaciones'
   when "informacion" like '%%TBILL%%' then 'Otros activos'
   when "tipoTitulo" like '%%ERNA%%' then 'ETF'
   when "tipoTitulo" like '%%Título%%euda%%' or "tipoTitulo" like '%%LETE%%' then 'Títulos Públicos'
   when "tipoTitulo" = 'Cheque de pago diferido' or "tipoTitulo" = 'Cheques de Pago Diferido' then 'CPD Directo'
   when "tipoTitulo" like '%%valado%%' then 'CPD Avalado'
   when "tipoTitulo" like '%%ECHEQ%%' then 'CPD Directo'
   when "tipoTitulo" is null and "informacion" like '%%Conciliar%%diferencias%%%%' then 'Conciliar diferencias recuperos'
   else "tipoTitulo"
   end as "tipoTitulo",
  "tipoOperacion",
  "operacion",
  "mercado",
  "segmento",
  "sesion",
  "parte",
  "especie",
  "concepto",
  "moneda",
  "cantidad",
  "valuacion",
  (case when "moneda" ~ 'USD.*' then round("cantidad"*(select tc from USD),2) else "valuacion" end) as "valuacionMEP"
from MOVS
where not("tipoTitulo" is null and "informacion" like '%%Conciliar%%diferencias%%recuperos%%' and "cantidad" = 0)
)"""

SQL_ARANCELES=SQL_ARANCELES_BASE+""", redu1 as (
select "tipoTitulo", TO_CHAR("concertacion", 'W' )::integer as "semana", TO_CHAR("concertacion", 'MM' )::integer as "mes", TO_CHAR("concertacion", 'YYYY' )::integer as "año", "moneda", "cantidad"
from movs2)
select "año","semana","mes","tipoTitulo","moneda",sum("cantidad")::real as "cantidad"
from redu1
group by "año","semana","mes","tipoTitulo","moneda"
order by 1,2,3,4
"""

SQL_ARANCELES_RANKING=SQL_ARANCELES_BASE+""", redu1 as (
select TO_CHAR("concertacion", 'W' )::integer as "semana",
       TO_CHAR("concertacion", 'MM' )::integer as "mes",
       TO_CHAR("concertacion", 'YYYY' )::integer as "año",
       "cuenta",
       "denominacionCuenta",
       "moneda",
       "cantidad"
from movs2)
select "semana","mes","año","cuenta","denominacionCuenta",
        round(sum("cantidad"*(case when "moneda"!='ARS' then %s else 1 end)),2)::real as "cantidad"
from redu1
group by "semana","mes","año","cuenta","denominacionCuenta"
order by 1,2,3,4 desc
"""

def query_aranceles(desde: date, hasta: date):
    conn: connection = _get_hg_connection()
    cur: cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(SQL_ARANCELES, (desde, hasta, desde, hasta))
    result = cur.fetchall()
    cur.close()
    conn.close()
    return result


def query_mep(hasta: date):
    conn: connection = _get_hg_connection()
    cur: cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(SQL_TC_MEP, (hasta,))
    result = cur.fetchone()
    cur.close()
    conn.close()
    return result["tc"]

def query_aranceles_ranking(desde: date, hasta: date, mep: float):
    conn: connection = _get_hg_connection()
    cur: cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(SQL_ARANCELES_RANKING, (desde, hasta, desde, hasta, mep))
    result = cur.fetchall()
    cur.close()
    conn.close()
    return result

def create_df_aranceles(desde: date, hasta: date) -> pd.DataFrame:
    result = query_aranceles(desde, hasta)
    mep = query_mep(hasta)
    df = pd.DataFrame(result)
    df["tc"] = df.apply(func=lambda row: mep, axis=1)
    df["cantidadUSD"] = df.apply(
        func=lambda row: (round(row["cantidad"] / mep, 2) if row["moneda"] == 'ARS' else row["cantidad"]), axis=1)
    df["cantidadARS"] = df.apply(
        func=lambda row: (round(row["cantidad"] * mep, 2) if row["moneda"] != 'ARS' else row["cantidad"]), axis=1)
    return df

def create_df_aranceles_ranking(desde: date, hasta: date) -> pd.DataFrame:
    mep = query_mep(hasta)
    result = query_aranceles_ranking(desde, hasta, mep)
    df = pd.DataFrame(result)
    return df

def create_report(desde: date, hasta: date):
    dfAranceles = create_df_aranceles(desde, hasta)
    dfArancelesRanking = create_df_aranceles_ranking(desde, hasta)

    workbook = load_workbook('TEMPLATE.xlsx')
    writer = pd.ExcelWriter("/tmp/deting.xlsx", engine='openpyxl')
    writer.book = workbook

    dfAranceles.to_excel(writer, "Datos Aranceles")
    dfArancelesRanking.to_excel(writer, "Datos Aranceles Ranking")

    writer.save()
