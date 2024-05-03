#!python3
import psycopg2
import psycopg2.extras
import redflagbpm
import pandas as pd
import datetime
from redflagbpm import PgUtils
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

bpm = redflagbpm.BPMService()


def _get_connection():
    # Manual connection, no config file
    conn = PgUtils.get_connection(bpm, 'FLW')
    conn.autocommit = True
    return conn


def consultar_pedidos_dia():

    sql_pos = """
           select *
from colocadoras_fci.cauciones
where monto !=0 or operado !=0
order by monto desc

            """
    conn = _get_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(sql_pos)
    data = cur.fetchall()
    cur.close()
    return pd.DataFrame(data)



def generar_excel():
    # Create a new Excel workbook
    wb = Workbook()

    # Remove the default "Sheet"
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # Create a new sheet for each DataFrame
    ws1 = wb.create_sheet(title='Cauciones')
    return wb, ws1


def write_sheet(sheet, dataframe):
    # Copiar el DataFrame a la hoja de Excel
    # Obtener el día actual
    fecha_actual = datetime.datetime.now()
    fecha = fecha_actual.strftime("%d/%m")
    hora = datetime.datetime.now().strftime("%H:%M")
    texto_fila = f"{fecha} {hora}"

    # Obtener la primera celda de la nueva fila
    cell = sheet.cell(row=1, column=1)

    # Establecer el valor de la celda con el texto del día
    cell.value = texto_fila

    # Combinar las celdas para que el texto abarque todas las columnas
    end_column = get_column_letter(sheet.max_column)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

    # Ajustar la alineación del texto
    cell.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False), 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Define bold font style
    bold_font = Font(bold=True)

    # Border
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Apply bold font to all cells and create a bold border
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = bold_font
            cell.border = border

    # Ajustar el ancho de las columnas
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width


def main():
    df_pedidos = consultar_pedidos_dia()
    #df_pedidos.columns = ['Nombre', 'Menú']

    wb, ws1 = generar_excel()
    write_sheet(ws1, df_pedidos)

    wb.save('/tmp/Pedidos.xlsx')


main()
#bpm.context.setJsonValue("_responseHeaders", "content-type", "application/vnd.ms-excel")
#bpm.context.setJsonValue("_responseHeaders", "resource", "/tmp/Pedidos.xlsx")
